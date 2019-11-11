#!/usr/bin/env python
# -*- coding:utf-8 -*-

import pymysql
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


def importExcelToMysql(cur, path):
    num = 1
    # 读取excel文件
    workbook = load_workbook(path)
    # 获得所有工作表的名字
    sheets = workbook.get_sheet_names()
    # 获得第一张表
    worksheet = workbook.get_sheet_by_name(sheets[0])

    # 将表中每一行数据读到sqlstr数组中
    # for row in worksheet.rows:
    #     if num == 1 or num == 2:
    #         num += 1
    #         continue
    #     sqlstr = []
    #     for cell in row:
    #         sqlstr.append(cell.value)
    #
    #     valuestr = [
    #         int(sqlstr[0]), str(sqlstr[1]), str(sqlstr[2]), str(sqlstr[3]),
    #         str(sqlstr[4]), str(sqlstr[5]), str(sqlstr[6]), str(sqlstr[7]),
    #         str(sqlstr[8]), str(sqlstr[9]), str(sqlstr[10]),
    #     ]
    #
    #     valuestr = tuple(valuestr)
    #     # 将每行数据存到数据库中
    #     cur.execute('use test')
    #     cur.execute(
    #         """insert into room(id, 管理区, 楼宇, 单元, 楼层, 房间代码, 房间类型, 建筑面积, 房间状态, 客户类别, 客户名称)
    #                         VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
    #         valuestr)

    """study"""
    wb = workbook
    ws = wb.get_sheet_by_name(sheets[0])
    # print('wb.active -->', wb.active)
    # print('wb.worksheets -->', wb.worksheets)
    # print('wb.encoding -->', wb.encoding)
    # print('wb.properties -->', wb.properties)
    # print('wb.sheetnames -->', wb.sheetnames)
    # print('wb.read_only -->', wb.read_only)

    """wb 方法"""
    # # 创建表 test
    # wb.create_sheet(title="测试", index=1)
    # wb.save(filename=path)
    # print('创建test')
    # 删除 test

    # print('wb.get_sheet_names()-->', wb.get_sheet_names())
    # ws = wb.active
    # print(ws)

    # copy 一个 sheet
    # ws1 = wb.copy_worksheet(from_worksheet=ws)
    # wb.save(filename=path)
    # print(ws1)
    # # 删除一个工作表
    # wb.remove(ws)
    # wb.save(filename=path)
    # print(wb.sheetnames)

    # # 合并单元格
    # ws.merge_cells('A2:D2')
    # ws.unmerge_cells('A2:D2')
    # wb.save(filename=path)
    # num = 1
    # for row in ws.rows:
    #     if num == 1:
    #         num = 0
    #         continue
    #     for cell in row:
    #         print(cell.value)
    #     break
    print(ws.rows)
    # print(ws.iter_rows())

    # print(ws.columns)
    # for col in ws.columns:
    #     for v in col:
    #         print(v)
    print(ws.iter_cols())
    for col in ws.iter_cols:
        for v in col:
            print(v)
# 输出数据库中的内容
def readTable(cursor):
    # 选择全部
    cursor.execute("select * from room")
    # 获得返回值， 返回多条记录， 若没有结果则返回
    results = cursor.fetchall()
    for i in range(0, results.__len__()):
        for j in range(0, 11):
            print(results[i][j], end='\t')
        print('\n')


def exportMysqlToExcel(cursor):
    wb = Workbook()
    # 创建一个活动的的sheet
    sheet = wb.active
    # 给活动的表赋值
    sheet.title = '导出数据'

    # 选择全部
    cursor.execute("use test")
    cursor.execute("select * from room")
    # 获得返回值， 返回多条记录， 若没有结果则返回
    results = cursor.fetchall()

    for i in range(0, results.__len__()):
        for j in range(0, 11):
            alpha = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G', 7: 'H', 8: 'I', 9: 'J', 10: 'K', }
            sheet["%s%d" % (alpha[j], i + 1)].value = results[i][j]

    # 保存数据到文件
    wb.save('./保存一个新的excel.xlsx')


def init(cur):
    # 新建一个database
    # cur.execute("drop database if exists test")
    # cur.execute("create database test")
    cur.execute("use test")

    sql = """
    CREATE TABLE  room(
        id int primary key,
        管理区 VARCHAR (20),
        楼宇 VARCHAR (20),
        单元 VARCHAR (20),
        楼层 int,
        房间代码 VARCHAR (20),
        房间类型 VARCHAR (20),
        建筑面积 float(6,2),
        房间状态  VARCHAR (20),
        客户类别  VARCHAR (20),
        客户名称  VARCHAR (20))
    """
    # 如果存在students这个表则删除
    cur.execute("drop table if exists room")
    # 创建表
    cur.execute(sql)


if __name__ == '__main__':
    """初始化数据库"""
    # 与数据库建立连接
    conn = pymysql.connect(host='127.0.0.1', user='root', password='root',
                           port=3306, charset='utf8')
    # 创建游标链接
    cur = conn.cursor()
    cur.execute('use test')
    # init(cur)
    # 将excel中的数据导入数据库中
    importExcelToMysql(cur, "/Users/henry/OneDrive/滨河物业管理/初始化数据/房间客户数据模板.xlsx")

    # 读取数据库
    # readTable(cur)

    # 导出数据
    # exportMysqlToExcel(cur)

    # 关闭游标链接
    cur.close()
    conn.commit()
    # 关闭数据库服务连接, 释放内存
    conn.close()
